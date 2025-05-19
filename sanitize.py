from __future__ import annotations

#!/usr/bin/env python3
"""
sanitize_excel_run.py
=====================
Excelファイル内の個人情報をサニタイズするスクリプト。SpaCy動作のためPython 3.12推奨。
1. VS Code でこのファイルを開く
2. CONFIG ブロックを書き換える
3. ツールバーの ▶Run  (または F5) を押す

すると:
    指定の Excel から PII を疑似値に置換し
    <出力ファイル> を保存 (置換マッピング JSON も任意で出力)

依存ライブラリや spaCy 日本語モデルは
最初に自動インストールされるので何もする必要はありません。
"""

# --------------------------------------------------------------------------- #
# CONFIG
# --------------------------------------------------------------------------- #
CONFIG = {
    "input_path": r"（インプットファイルパス）",
    "sheets": ["（インプットシート名）"],
    "output_path": r"（アウトプットファイル名）",
    "mapping_json_path": None,
}
# --------------------------------------------------------------------------- #

import dataclasses as _dc
import importlib
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Tuple

_PKGS = ("openpyxl", "faker", "spacy")

def _ensure_pkg(name: str) -> None:
    try:
        importlib.import_module(name)
    except ModuleNotFoundError:
        print(f"[INFO] Installing '{name}' ...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--user", name],
            stdout=subprocess.DEVNULL,
        )

for _p in _PKGS:
    _ensure_pkg(_p)

import spacy
try:
    spacy.load("ja_core_news_lg")
except OSError:
    print("[INFO] Downloading spaCy model 'ja_core_news_lg' ...")
    import spacy.cli
    spacy.cli.download("ja_core_news_lg")

import faker
import openpyxl
_NLP = spacy.load("ja_core_news_lg")

import unicodedata

# ---------------------- 正規表現パターン ------------------------------ #
_REGEX_PATTERNS: dict[str, str] = {
    "PHONE": r"""(?<!\d)(?:
         0(?:70|80|90)(?:[-‐‑]?\d{4}){2}
       | 0\d{1,4}[-‐‑]?\d{1,4}[-‐‑]?\d{3,4}
       | 0(?:70|80|90)\d{8}
       | 0\d{9,10}
    )(?!\d)""",

    "EMAIL": r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",

    "ZIP":   r"(?:〒\s*)?\b\d{3}[-‐‑]\d{4}\b",

    "IP": r"\b(?:\d{1,3}\.){3}\d{1,3}\b",

    "CREDIT_CARD": r"\b(?:\d[ -]*?){13,16}\b",

    "ADDR": r"""(?<![一-龥0-9])(?:
         (?:(?:北海道|(?:京都|大阪)府|.{2,3}県)?[一-龥]{1,8}(?:市|区|町|村)[一-龥]{1,8}[\-‑‐]?\d{1,3}(?:丁目?)?)
       | [一-龥]{2,10}\d{1,3}(?:[一-龥])?
    )""",

    # マイナンバー関連（文脈付き）
    "MY_NUMBER": r"(?<!\d)(?:(?:マイナンバー[:：]?\s*)?)\d{12}(?!\d)",

    "PIN_USER": r"(?i)(?:利用者証明用暗証番号[:：]?\s*)\d{4}",

    "PIN_SIGNATURE": r"(?i)(?:署名用パスワード[:：]?\s*)[A-Za-z0-9]{6,16}",

    "PIN_KENMEN": r"(?i)(?:券面補助AP[:：]?\s*)\d{4}",

    "PIN_JUMIN": r"(?i)(?:住民基本台帳用暗証番号[:：]?\s*)\d{4}",
}

_SPACY_LABELS = {"PERSON", "GPE", "LOC", "DATE", "ORG", "ADDR"}
_FAKE = faker.Faker("ja_JP")

@_dc.dataclass(slots=True)
class Replacement:
    span: Tuple[int, int]
    label: str
    fake: str

def _fake(label: str) -> str:
    fixed = {
        "PERSON":       "namexxx",
        "EMAIL":        "mailxxx",
        "PHONE":        "phonexxx",
        "ZIP":          "zipxxx",
        "IP":           "ipxxx",
        "CREDIT_CARD":  "cardxxx",
        "GPE":          "prefxxx",
        "LOC":          "cityxxx",
        "DATE":         "datexxx",
        "ORG":          "orgxxx",
        "ADDR":         "addrxxx",
        "MY_NUMBER":    "mynumberxxx",
        "PIN_USER":     "pinuserxxx",
        "PIN_SIGNATURE":"pinsigxxx",
        "PIN_KENMEN":   "pinkmxxx",
        "PIN_JUMIN":    "pinjuminxxx",
    }
    return fixed.get(label, f"<{label}hogehoge>")

def _detect_regex(txt: str) -> List[Replacement]:
    flags = re.I | re.UNICODE | re.VERBOSE
    out: List[Replacement] = []
    for lab, pat in _REGEX_PATTERNS.items():
        for m in re.finditer(pat, txt, flags=flags):
            out.append(Replacement(m.span(), lab, _fake(lab)))
    return out

def _detect_spacy(txt: str) -> List[Replacement]:
    doc = _NLP(txt)
    return [
        Replacement((e.start_char, e.end_char), e.label_, _fake(e.label_))
        for e in doc.ents if e.label_ in _SPACY_LABELS
    ]

def sanitize_text(txt: str) -> Tuple[str, Dict[str, str]]:
    repls = _detect_regex(txt) + _detect_spacy(txt)
    repls.sort(key=lambda r: (r.span[0], -r.span[1]))

    merged, end = [], -1
    for r in repls:
        if r.span[0] >= end:
            merged.append(r)
            end = r.span[1]

    mapping, out = {}, txt
    for r in reversed(merged):
        orig = out[r.span[0]:r.span[1]]
        out = out[:r.span[0]] + r.fake + out[r.span[1]:]
        mapping[orig] = r.fake
    return out, mapping

def sanitize_workbook(in_path: Path, sheets: List[str] | None):
    wb = openpyxl.load_workbook(in_path, data_only=False, keep_links=True)
    targets = sheets or wb.sheetnames
    all_map = {}

    for name in targets:
        if name not in wb.sheetnames:
            print(f"[WARN] Sheet '{name}' not found; skip.")
            continue
        ws, sheet_map = wb[name], {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    san, mp = sanitize_text(cell.value)
                    if mp:
                        cell.value = san
                        sheet_map.update(mp)
        all_map[name] = sheet_map
        print(f"[INFO] {name}: {len(sheet_map)} replacements")
    return wb, all_map

def main():
    in_path = Path(CONFIG["input_path"]).expanduser()
    out_path = (
        Path(CONFIG["output_path"])
        if CONFIG["output_path"]
        else in_path.with_stem(in_path.stem + "_sanitized")
    )
    sheets = CONFIG["sheets"]
    map_json = CONFIG["mapping_json_path"]

    if not in_path.is_file():
        sys.exit(f"Input not found: {in_path}")

    wb, mapping = sanitize_workbook(in_path, sheets)
    wb.save(out_path)
    print(f"[DONE] Saved: {out_path}")

    if map_json:
        Path(map_json).write_text(
            json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"[INFO] Mapping JSON: {map_json}")

if __name__ == "__main__":
    if os.name == "nt":
        os.environ.setdefault("PYTHONUTF8", "1")
    main()
